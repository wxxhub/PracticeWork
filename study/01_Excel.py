
# 【1】==导包openpyxl
from openpyxl import Workbook;
# 【2】==创建一个Excel表格
wb = Workbook();
# 【4】==进行表格操作
# 1.默认创建的表格会有一个默认的选项卡：Sheet
# 2.添加选项卡
wb.create_sheet("第一个选项卡",0);
wb.create_sheet("第二个选项卡",1);
# 3.删除选项卡
wb.remove_sheet(wb["Sheet"]);
# 4.获取要操控的选项卡
# 获取当前活跃状态的选项卡-正在操作或最新打开的选项卡
ws = wb.active;
print(ws)
# 5.根据选项卡名称选择要操控的选项卡
ws1=wb["第二个选项卡"];
print(ws1)
# 6.赋值操作
ws1["A1"] = "姓名";
# 7.根据行和列的形式操作表格
ws1.cell(row=4,column=4,value="二〇一九年三月六日");
# 8.整行输入
ws.append(["姓名","年龄","性别","身份"]);
ws.append(["狄仁杰","18","男","射手"]);
# 9.循环添加
# 创建一个二位数组
array_01 = [["姓名","年龄","性别","身份"],["狄仁杰","18","男","射手"],["狄仁杰","18","男","射手"]]
for i in range(0,len(array_01)):
    ws1.append(array_01[i]);
    pass
ws.add_image("")




# 【3】==存储表格
wb.save("2019年3月6日.xlsx")
print("表格�