from openpyxl import  Workbook;
wb = Workbook();

wb.create_sheet("第一个sheet",0);
wb.create_sheet("第二个sheet",1);
wb.remove_sheet(wb["Sheet"]);


wb.save("2019年3月6日.xlsx");
print("保存成功！");